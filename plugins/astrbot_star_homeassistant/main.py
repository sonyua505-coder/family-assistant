"""
家庭信息助理 · AstrBot Star 插件（M7，全功能接入）

接入层：让 QQ 机器人（鸡煲大人2.0）通过 LLM 工具调用 homeassistant API。
  - 身份：get_my_identity / create_person / bind_person / unbind_identity /
    set_primary_identity，身份头由插件注入（x-platform=qq,
    x-openid=消息发送者 openid，见 src/lib/identity.ts）。
  - 账本：create_account / list_my_accounts / join_account。
  - 记账：add_bill / add_bills(批量) / update_bill / delete_bill / settle_bill /
    bill_stats(统计图) / export_bills(查询/导出)（含 AA 分摊）。
  - 待办：add_task / list_tasks / complete_task / undo_task / update_task / delete_task。
  - 工作账单（装修安装门，2026-08-18）：add_work_client / set_work_price / add_work_bill /
    list_work_bills / get_work_bill / settle_work_bill / recalc_work_bills / export_work_bills 等。
  - 订阅/新闻：subscribe_source / list_subscriptions / unsubscribe / query_news。
  - 抓取：fetch_source（SSRF 白名单由后端把关）。
  - 账本链接：create_ledger_link / list_ledger_links / revoke_ledger_link。
  - outbox：后台轮询 /api/v1/outbox/pending?channel=qq，经
    context.send_message(umo, ...) 主动推送到 QQ（C2C 无需 msg_id，
    见 qqofficial 适配器 _send_by_session_common），并回报 sent/failed。

配置（WebUI 插件面板 / _conf_schema.json）：
  api_base      家庭信息助理 API 地址（默认 http://homeassistant-api:3000，走共享网络）
  api_key       X_API_KEY（/opt/homeassistant/.env，勿外泄）
  poll_interval outbox 轮询间隔秒数
"""
import asyncio
import csv
import io
import os
import re
import shutil
import uuid
from datetime import date as _date
from datetime import datetime
from pathlib import Path
from urllib.parse import urlencode

import httpx

from astrbot.api import logger
from astrbot.api.event import AstrMessageEvent, filter
from astrbot.api.message_components import File, Reply
from astrbot.api.star import Context, Star
from astrbot.core.message.message_event_result import MessageChain
from astrbot.core.utils.astrbot_path import get_astrbot_temp_path, get_astrbot_workspaces_path
from astrbot.core.workspace import normalize_umo_for_workspace

DEFAULT_API_BASE = "http://homeassistant-api:3000"
DEFAULT_POLL_INTERVAL = 15


class HomeAssistantStar(Star):
    def __init__(self, context: Context, config: dict = None):
        super().__init__(context)
        self.config = config or {}
        self.api_base = str(self.config.get("api_base") or DEFAULT_API_BASE).rstrip("/")
        self.api_key = str(self.config.get("api_key") or "").strip()
        try:
            self.poll_interval = max(5, int(self.config.get("poll_interval") or DEFAULT_POLL_INTERVAL))
        except (TypeError, ValueError):
            self.poll_interval = DEFAULT_POLL_INTERVAL
        # openid -> unified_msg_origin（从每个入站消息捕获），供 outbox 主动推送定位会话
        self._session_map: dict[str, str] = {}
        # qq_official 适配器的平台 id（UMO 首段），用于无注册表命中时构造 C2C 会话
        self._platform_id: str | None = None
        self._outbox_task: asyncio.Task | None = None

    # ────────────────────────── 生命周期 ──────────────────────────

    async def initialize(self) -> None:
        """插件激活时启动 outbox 轮询。"""
        self._outbox_task = asyncio.create_task(self._outbox_loop())
        logger.info(
            f"[HomeAssistant] 插件已激活 api_base={self.api_base} "
            f"poll_interval={self.poll_interval}s"
        )

    async def terminate(self) -> None:
        if self._outbox_task:
            self._outbox_task.cancel()
            try:
                await self._outbox_task
            except (asyncio.CancelledError, Exception):
                pass
            self._outbox_task = None

    # ────────────────────────── HTTP 帮助 ──────────────────────────

    async def _api(self, method: str, path: str, *, identity=None, json_body=None) -> dict:
        """调 homeassistant API；identity=(platform, openid) 时注入身份头。"""
        headers = {"Authorization": f"Bearer {self.api_key}"}
        if identity:
            headers["x-platform"] = identity[0]
            headers["x-openid"] = identity[1]
        url = f"{self.api_base}/{path.lstrip('/')}"
        try:
            async with httpx.AsyncClient(timeout=10) as client:
                resp = await client.request(method, url, json=json_body, headers=headers)
            logger.debug(f"[HomeAssistant] {method} {path} -> {resp.status_code}")
            if resp.status_code >= 400:
                return {"ok": False, "http": resp.status_code, "body": resp.text[:500]}
            try:
                return resp.json()
            except ValueError:
                return {"ok": False, "http": resp.status_code, "body": resp.text[:500]}
        except Exception as e:  # 网络/超时等，统一返回失败字典，让工具层提示
            logger.error(f"[HomeAssistant] 请求失败 {method} {path}: {e}")
            return {"ok": False, "error": str(e)}

    @staticmethod
    def _identity(event: AstrMessageEvent):
        return ("qq", event.get_sender_id())

    @staticmethod
    def _err_msg(data: dict) -> str:
        """把后端失败响应转成对用户友好的一句话。"""
        body = data.get("body") or data.get("error") or str(data)
        if "ACCOUNT_AMBIGUOUS" in str(body):
            return "你有多个账本，请先 list_my_accounts 查看账本 id，再带上 account_id 重试。"
        return f"操作失败：{body}"

    async def _api_raw(self, method: str, path: str, *, identity=None, json_body=None):
        """调 homeassistant API 并返回 (status_code, body_text)；网络异常返回失败 dict。

        _api 只支持 JSON 响应；导出 CSV 是纯文本，需要走这个原始接口。
        """
        headers = {"Authorization": f"Bearer {self.api_key}"}
        if identity:
            headers["x-platform"] = identity[0]
            headers["x-openid"] = identity[1]
        url = f"{self.api_base}/{path.lstrip('/')}"
        try:
            async with httpx.AsyncClient(timeout=60) as client:  # 导出可能较慢，放宽超时
                resp = await client.request(method, url, json=json_body, headers=headers)
            return resp.status_code, resp.text
        except Exception as e:
            logger.error(f"[HomeAssistant] 请求失败 {method} {path}: {e}")
            return {"ok": False, "error": str(e)}

    def _workspace_root(self, event: AstrMessageEvent) -> Path:
        """当前会话工作区根目录（resolve 后），按 unified_msg_origin 隔离。"""
        root = Path(get_astrbot_workspaces_path()) / normalize_umo_for_workspace(
            event.unified_msg_origin
        )
        return root.resolve()

    def _resolve_workspace_rel(self, event: AstrMessageEvent, relative_path: str) -> Path:
        """把相对路径补全到会话工作区，并校验不越界（容器 root 无沙箱，必须插件自查）。"""
        rel = str(relative_path).strip().replace("\\", "/")
        if not rel or rel.startswith("/") or Path(rel).is_absolute():
            raise ValueError("必须是工作区内的相对路径")
        root = self._workspace_root(event)
        target = (root / rel).resolve()
        if target == root or not target.is_relative_to(root):
            raise ValueError(f"路径越界：{relative_path}")
        return target

    def _ensure_matplotlib(self) -> bool:
        """懒加载 matplotlib（Agg 后端 + 中文字体注册），失败返回 False。

        容器无显示、装失败不影响插件加载，故绝不在模块顶层 import。
        """
        if getattr(self, "_mpl_ok", None) is not None:
            return self._mpl_ok
        try:
            import matplotlib

            matplotlib.use("Agg")  # 容器无显示，必须用非交互后端
            from matplotlib import font_manager, pyplot as plt

            self._plt = plt
            font_path = os.path.join(
                os.path.dirname(os.path.abspath(__file__)), "fonts", "wqy-microhei.ttc"
            )
            if os.path.exists(font_path):
                font_manager.fontManager.addfont(font_path)
                family = font_manager.FontProperties(fname=font_path).get_name()
                plt.rcParams["font.family"] = [family, "DejaVu Sans"]
            else:
                logger.warning("[HomeAssistant] 中文字体缺失，图表中文可能显示为方块")
            plt.rcParams["axes.unicode_minus"] = False  # 修复负号显示
            self._mpl_ok = True
            return True
        except Exception as e:
            logger.error(f"[HomeAssistant] matplotlib 初始化失败: {e}")
            self._mpl_ok = False
            return False

    def _render_stats_png(self, stats: dict, period: str, out_path: str) -> None:
        """渲染统计图（CPU 密集，经 asyncio.to_thread 调用；金额单位：元，仅供用户看图）。

        一张 PNG 两个 subplot：左=支出分类构成（横向条形，Top6+其他），右=收支趋势（分组柱状）。
        qqofficial 一条消息只支持一个 Image，故合成单张。
        """
        import numpy as np

        plt = self._plt
        by_cat = [c for c in (stats.get("by_category") or []) if c.get("amount")]
        by_cat.sort(key=lambda c: c["amount"], reverse=True)
        trend = stats.get("trend") or []

        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 5), dpi=150)
        fig.suptitle(f"家庭信息助理 · 收支统计 {period}", fontsize=13)

        # 左：支出分类构成 —— 横向条形图（part-to-whole 用条形，小屏标签可读）
        if by_cat:
            top = by_cat[:6]
            labels = [c["category"] for c in top]
            vals = [c["amount"] / 100 for c in top]
            rest = sum(c["amount"] for c in by_cat[6:]) / 100
            if rest > 0:
                labels.append("其他")
                vals.append(rest)
            y = np.arange(len(vals))[::-1]
            ax1.barh(y, vals, color="#2a78d6", height=0.62)
            ax1.set_yticks(y, labels)
            ax1.set_title("支出分类构成（元）")
            for yi, v in zip(y, vals):
                ax1.text(v, yi, f"{v:,.0f}", va="center", ha="left", fontsize=9)
            ax1.tick_params(axis="x", labelsize=8)
        else:
            ax1.text(0.5, 0.5, "暂无分类数据", ha="center", va="center", transform=ax1.transAxes)

        # 右：收支趋势 —— 分组柱状（收入蓝 / 支出橙）
        if trend:
            months = [t.get("month", "") for t in trend]
            inc = [t.get("income", 0) / 100 for t in trend]
            exp = [t.get("expense", 0) / 100 for t in trend]
            x = np.arange(len(months))
            w = 0.36
            b1 = ax2.bar(x - w / 2, inc, w, color="#2a78d6", label="收入")
            b2 = ax2.bar(x + w / 2, exp, w, color="#eb6834", label="支出")
            ax2.set_xticks(x, months, rotation=30, ha="right", fontsize=8)
            ax2.set_ylabel("元")
            ax2.legend(frameon=False)
            ax2.bar_label(b1, fontsize=8, padding=2)
            ax2.bar_label(b2, fontsize=8, padding=2)
        else:
            ax2.text(0.5, 0.5, "暂无趋势数据", ha="center", va="center", transform=ax2.transAxes)

        fig.tight_layout(rect=[0, 0, 1, 0.95])
        fig.savefig(out_path, bbox_inches="tight", facecolor="white")
        plt.close(fig)

    def _render_work_stats_png(self, stats: dict, period: str, out_path: str) -> None:
        """渲染工作账单统计图（经 asyncio.to_thread；金额单位：元）。

        左=按委托方欠款（横向条形，Top6+其他）；右=按月份应收/已收（分组柱状）。
        """
        import numpy as np

        plt = self._plt
        by_client = sorted(
            [c for c in (stats.get("by_client") or []) if c.get("owed")],
            key=lambda c: c["owed"],
            reverse=True,
        )
        by_month = stats.get("by_month") or []

        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 5), dpi=150)
        fig.suptitle(f"家庭信息助理 · 工作账单统计 {period}", fontsize=13)

        # 左：按委托方欠款 —— 横向条形
        if by_client:
            top = by_client[:6]
            labels = [c["client_name"] for c in top]
            vals = [c["owed"] / 100 for c in top]
            rest = sum(c["owed"] for c in by_client[6:]) / 100
            if rest > 0:
                labels.append("其他")
                vals.append(rest)
            y = np.arange(len(vals))[::-1]
            ax1.barh(y, vals, color="#eb6834", height=0.62)
            ax1.set_yticks(y, labels)
            ax1.set_title("按委托方欠款（元）")
            for yi, v in zip(y, vals):
                ax1.text(v, yi, f"{v:,.0f}", va="center", ha="left", fontsize=9)
            ax1.tick_params(axis="x", labelsize=8)
        else:
            ax1.text(0.5, 0.5, "暂无委托方数据", ha="center", va="center", transform=ax1.transAxes)

        # 右：按月份应收/已收 —— 分组柱状（应收蓝 / 已收绿）
        if by_month:
            months = [m.get("month", "") for m in by_month]
            recv = [m.get("receivable", 0) / 100 for m in by_month]
            paid = [m.get("paid", 0) / 100 for m in by_month]
            x = np.arange(len(months))
            w = 0.36
            b1 = ax2.bar(x - w / 2, recv, w, color="#2a78d6", label="应收")
            b2 = ax2.bar(x + w / 2, paid, w, color="#16a34a", label="已收")
            ax2.set_xticks(x, months, rotation=30, ha="right", fontsize=8)
            ax2.set_ylabel("元")
            ax2.legend(frameon=False)
            ax2.bar_label(b1, fontsize=8, padding=2)
            ax2.bar_label(b2, fontsize=8, padding=2)
        else:
            ax2.text(0.5, 0.5, "暂无月份数据", ha="center", va="center", transform=ax2.transAxes)

        fig.tight_layout(rect=[0, 0, 1, 0.95])
        fig.savefig(out_path, bbox_inches="tight", facecolor="white")
        plt.close(fig)

    # ────────────────────────── 会话注册 ──────────────────────────

    @filter.regex(r".*")
    async def _capture_session(self, event: AstrMessageEvent):
        """每个入站消息记录 openid -> unified_msg_origin，供 outbox 定位会话。"""
        try:
            sender = event.get_sender_id()
            umo = event.unified_msg_origin
            if sender and umo:
                self._session_map[sender] = umo
                if self._platform_id is None:
                    self._platform_id = umo.split(":", 1)[0]
        except Exception as e:
            logger.debug(f"[HomeAssistant] 会话捕获跳过: {e}")

    # ────────────────────────── LLM 工具 ──────────────────────────

    @filter.llm_tool(name="get_my_identity")
    async def get_my_identity(self, event: AstrMessageEvent) -> str:
        """查询当前 QQ 用户是否已登记身份及其账户。

        Args:
        """
        data = await self._api("GET", "api/v1/identity", identity=self._identity(event))
        if not data.get("ok", True):
            return f"查询失败：{data.get('body') or data.get('error') or data}"
        if data.get("bound"):
            p = data.get("person") or {}
            name = p.get("display_name") or p.get("nickname") or "（无名字）"
            accounts = data.get("accounts") or []
            acc = "，".join(
                f"{a.get('name')}(#{a.get('id')})" for a in accounts
            ) or "无（请先创建账本）"
            return f"已登记：{name}。可用账本：{acc}"
        return "该用户尚未登记身份。请先询问用户希望怎么称呼，然后调用 create_person 登记。"

    @filter.llm_tool(name="create_person")
    async def create_person(self, event: AstrMessageEvent, display_name: str) -> str:
        """为当前 QQ 用户创建/登记身份（首次使用引导，幂等）。

        Args:
            display_name(string): 用户希望被称呼的名字。
        """
        data = await self._api(
            "POST", "api/v1/persons",
            identity=self._identity(event), json_body={"display_name": display_name},
        )
        if not data.get("ok", True):
            return f"登记失败：{data.get('body') or data.get('error') or data}"
        p = data.get("person") or {}
        name = p.get("display_name") or p.get("nickname") or display_name
        if data.get("already_bound"):
            return f"该账号已登记过身份（{name}），无需重复创建。"
        return (
            f"登记成功！以后称呼你为「{name}」。请先创建你的记账账本"
            f"（个人账本），创建后就可以记账了。"
        )

    @filter.llm_tool(name="create_account")
    async def create_account(
        self,
        event: AstrMessageEvent,
        type: str = "personal",
        name: str = None,
    ) -> str:
        """为当前用户创建一个记账账户（新用户首次记账前必须已有账户）。

        Args:
            type(string): personal（个人账本）或 family（家庭账本），默认 personal。
            name(string): 账本名称，如「我的账本」「我家」，可选。
        """
        body = {"type": type}
        if name:
            body["name"] = name
        data = await self._api(
            "POST", "api/v1/accounts", identity=self._identity(event), json_body=body,
        )
        if not data.get("ok", True):
            return f"创建账本失败：{data.get('body') or data.get('error') or data}"
        a = data.get("account") or {}
        return (
            f"账本已创建：{a.get('name')}(#{a.get('id')})，类型 {a.get('type')}。"
            f"现在可以记账了，例如「记一笔：午饭 30 元」。"
        )

    @filter.llm_tool(name="list_my_accounts")
    async def list_my_accounts(self, event: AstrMessageEvent) -> str:
        """列出当前用户可用的记账账户（有多个账本时记账需指定 account_id）。"""
        data = await self._api("GET", "api/v1/accounts", identity=self._identity(event))
        if not data.get("ok", True):
            return f"查询失败：{data.get('body') or data.get('error') or data}"
        accounts = data.get("accounts") or []
        if not accounts:
            return "你还没有记账账户，请先调用 create_account 创建一个。"
        lines = [
            f"{a.get('name')}(#{a.get('id')})，类型 {a.get('type')}"
            for a in accounts
        ]
        return "你的账本：\n" + "\n".join(lines)

    @filter.llm_tool(name="add_bill")
    async def add_bill(
        self,
        event: AstrMessageEvent,
        type: str,
        amount: float,
        category: str = None,
        note: str = None,
        participants: list = None,
        account_id: int = None,
    ) -> str:
        """记录一笔账（单笔直记；AA/平摊时务必把参与者列出，之后可再结算）。

        Args:
            type(string): income 或 expense（收入/支出）。
            amount(number): 金额（单位：分，1 元 = 100 分）。
            category(string): 分类，如 餐饮/交通/购物。
            note(string): 备注。
            participants(array): AA 分摊的参与者名字列表，如 ["张三","李四"]。用户说「AA」「平摊」「垫付」时先问清参与者有谁再传；只有一人/无分摊则不传。
            account_id(number): 记账账户 id，用户有多个账本时必须指定。
        """
        body = {"type": type, "amount": int(amount)}
        if category:
            body["category"] = category
        if note:
            body["note"] = note
        if participants:
            body["participants"] = self._normalize_participants(participants)
        if account_id:
            body["account_id"] = account_id
        data = await self._api(
            "POST", "api/v1/bills", identity=self._identity(event), json_body=body,
        )
        if not data.get("ok", True):
            return f"记账失败：{data.get('body') or data.get('error') or data}"
        b = data.get("bill") or {}
        kind = "支出" if b.get("type") == "expense" else "收入"
        aa = "（AA 待结算）" if b.get("status") == "pending" else ""
        return (
            f"已记录#{b.get('id')}：{kind} {b.get('amount')}分 "
            f"（{b.get('category') or '未分类'}）{aa}备注：{b.get('note') or '无'}"
        )

    @filter.llm_tool(name="settle_bill")
    async def settle_bill(
        self,
        event: AstrMessageEvent,
        bill_id: int,
        participant_name: str = None,
        all: bool = False,
    ) -> str:
        """结算一笔 AA 账单（标记某参与人已付款，或一次性全部结清）。

        Args:
            bill_id(number): 账单 id（来自记账或查账结果的 #id）。
            participant_name(string): 要结算的参与人名字（可选）。
            all(boolean): 是否全部结清（可选，默认 false）。
        """
        body = {}
        if participant_name:
            body["participant_name"] = participant_name
        if all:
            body["all"] = True
        data = await self._api(
            "POST", f"api/v1/bills/{bill_id}/settle",
            identity=self._identity(event), json_body=body,
        )
        if not data.get("ok", True):
            return f"结算失败：{data.get('body') or data.get('error') or data}"
        b = data.get("bill") or {}
        state = "已全部结清" if b.get("status") == "settled" else "仍待结算"
        return f"账单#{b.get('id')}：{state}。参与人：{b.get('participants')}"

    @filter.llm_tool(name="update_bill")
    async def update_bill(
        self,
        event: AstrMessageEvent,
        bill_id: int,
        type: str = None,
        amount: float = None,
        category: str = None,
        note: str = None,
        participants: list = None,
        occurred_at: str = None,
    ) -> str:
        """修改一笔已记录的账单（只需传要改的字段，其他不变）。

        Args:
            bill_id(number): 账单 id（来自记账或查账结果的 #id）。
            type(string): income 或 expense（可选）。
            amount(number): 金额（分）（可选）。
            category(string): 分类（可选）。
            note(string): 备注（可选）。
            participants(array): AA 参与者名单（可选）。
            occurred_at(string): 发生日期，YYYY-MM-DD 或 YYYY-MM-DD HH:MM:SS（可选，修正导入的日期错误）。
        """
        body = {}
        if type:
            body["type"] = type
        if amount is not None:
            body["amount"] = int(amount)
        if category:
            body["category"] = category
        if note:
            body["note"] = note
        if participants:
            body["participants"] = self._normalize_participants(participants)
        if occurred_at:
            body["occurred_at"] = occurred_at
        data = await self._api(
            "PATCH", f"api/v1/bills/{bill_id}",
            identity=self._identity(event), json_body=body,
        )
        if not data.get("ok", True):
            return f"修改失败：{data.get('body') or data.get('error') or data}"
        b = data.get("bill") or {}
        kind = "支出" if b.get("type") == "expense" else "收入"
        return (
            f"已修改#{b.get('id')}：{kind} {b.get('amount')}分 "
            f"（{b.get('category') or '未分类'}）备注：{b.get('note') or '无'}"
        )

    @filter.llm_tool(name="delete_bill")
    async def delete_bill(self, event: AstrMessageEvent, bill_id: int) -> str:
        """删除一笔账单（软删除，进回收站，可恢复）。

        Args:
            bill_id(number): 账单 id。
        """
        data = await self._api(
            "DELETE", f"api/v1/bills/{bill_id}", identity=self._identity(event),
        )
        if not data.get("ok", True):
            return f"删除失败：{data.get('body') or data.get('error') or data}"
        return f"已删除账单#{bill_id}（软删除；回收站查看/恢复请在记账网页「回收站」页操作）。"

    @staticmethod
    def _normalize_participants(participants: list) -> list:
        """把 LLM 可能给出的 ["张三","李四"] 或 [{"name":"张三"}] 归一成 API 需要的 [{"name":...}]。"""
        out = []
        for p in participants:
            if isinstance(p, str) and p.strip():
                out.append({"name": p.strip()})
            elif isinstance(p, dict) and p.get("name"):
                out.append({"name": p["name"]})
        return out

    # ────────────────────────── 批量/变动 ──────────────────────────

    @filter.llm_tool(name="add_bills")
    async def add_bills(
        self,
        event: AstrMessageEvent,
        bills: list,
        account_id: int = None,
    ) -> str:
        """批量补录多笔账单（先向用户复述清单、确认无误后再调用本工具入库）。

        Args:
            bills(array): 账单数组，每笔含 {type, amount, category?, note?, occurred_at?, participants?}。
            account_id(number): 记账账户 id（可选）。
        """
        body = {"bills": []}
        for b in bills or []:
            if not isinstance(b, dict):
                continue
            item = {"type": b.get("type"), "amount": int(b.get("amount") or 0)}
            if b.get("category"):
                item["category"] = b["category"]
            if b.get("note"):
                item["note"] = b["note"]
            if b.get("occurred_at"):
                item["occurred_at"] = b["occurred_at"]
            if b.get("participants"):
                item["participants"] = self._normalize_participants(b["participants"])
            body["bills"].append(item)
        if account_id:
            body["account_id"] = account_id
        data = await self._api(
            "POST", "api/v1/bills/batch", identity=self._identity(event), json_body=body,
        )
        if not data.get("ok", True):
            return f"批量录入失败：{data.get('body') or data.get('error') or data}"
        return f"已批量录入 {data.get('count')} 笔，ids={data.get('bill_ids')}"

    # ────────────────────────── 待办事项 ──────────────────────────

    @filter.llm_tool(name="add_task")
    async def add_task(
        self,
        event: AstrMessageEvent,
        content: str,
        category: str = None,
        remind_at: str = None,
        account_id: int = None,
    ) -> str:
        """添加一条家庭待办事项。

        Args:
            content(string): 事项内容。
            category(string): 分类（可选）。
            remind_at(string): 提醒时间，格式 YYYY-MM-DD HH:MM（可选）。
            account_id(number): 归属账户 id（可选）。
        """
        body = {"content": content}
        if category:
            body["category"] = category
        if remind_at:
            body["remind_at"] = remind_at
        if account_id:
            body["account_id"] = account_id
        data = await self._api(
            "POST", "api/v1/tasks", identity=self._identity(event), json_body=body,
        )
        if not data.get("ok", True):
            return f"添加失败：{data.get('body') or data.get('error') or data}"
        t = data.get("task") or {}
        return f"已添加事项#{t.get('id')}：{t.get('content')}（{t.get('category') or '未分类'}）"

    @filter.llm_tool(name="list_tasks")
    async def list_tasks(
        self,
        event: AstrMessageEvent,
        is_done: bool = None,
        q: str = None,
        category: str = None,
        from_date: str = None,
        to_date: str = None,
        page: int = None,
        page_size: int = None,
        account_id: int = None,
    ) -> str:
        """列出家庭待办事项，支持关键词/分类/日期区间/分页。默认只看未完成（需看全部请传 is_done）。

        Args:
            is_done(boolean): 状态筛选，true=已完成，false=未完成；不传则默认只看未完成（可选）。
            q(string): 内容关键词，空格分隔多个词须全部命中（可选）。
            category(string): 分类筛选，如 购物（可选）。
            from_date(string): 起始创建日期 YYYY-MM-DD（可选）。
            to_date(string): 结束创建日期 YYYY-MM-DD（可选）。
            page(number): 页码，从 1 开始（可选）。
            page_size(number): 每页条数，默认 50，最大 200（可选）。
            account_id(number): 归属账户 id（有多个账本时必须指定）。
        """
        params = {}
        if is_done is not None:
            params["is_done"] = "1" if is_done else "0"
        if q:
            params["q"] = q
        if category:
            params["category"] = category
        if from_date:
            params["from"] = from_date
        if to_date:
            params["to"] = to_date
        if page:
            params["page"] = page
        if page_size:
            params["page_size"] = page_size
        if account_id:
            params["account_id"] = account_id
        qs = "&".join(f"{k}={v}" for k, v in params.items())
        data = await self._api(
            "GET", f"api/v1/tasks?{qs}", identity=self._identity(event),
        )
        if not data.get("ok", True):
            return self._err_msg(data)
        items = data.get("items") or []
        total = data.get("total") or len(items)
        applied = data.get("applied") or {}
        if not items:
            return "没有符合条件的待办。" + (f"\n已生效筛选：{applied}" if applied else "")
        lines = [f"共 {total} 条待办（显示前 {len(items)} 条）："]
        for it in items[:20]:
            done = "✅" if it.get("is_done") else "⬜"
            remind = f" 提醒:{it.get('remind_at')}" if it.get("remind_at") else ""
            lines.append(f"{done} #{it.get('id')} {it.get('content')}{remind}")
        if total > len(items):
            lines.append(f"\n…还有 {total - len(items)} 条未显示，可缩小条件或用 export_tasks 导出。")
        if applied:
            lines.append(f"已生效筛选：{applied}")
        return "\n".join(lines)

    @filter.llm_tool(name="complete_task")
    async def complete_task(self, event: AstrMessageEvent, task_id: int) -> str:
        """标记一条待办已完成。

        Args:
            task_id(number): 事项 id。
        """
        data = await self._api(
            "POST", f"api/v1/tasks/{task_id}/done", identity=self._identity(event),
        )
        if not data.get("ok", True):
            return f"操作失败：{data.get('body') or data.get('error') or data}"
        return f"已完成事项#{task_id}。"

    @filter.llm_tool(name="undo_task")
    async def undo_task(self, event: AstrMessageEvent, task_id: int) -> str:
        """把一条已完成待办改回未完成（误点完成时可反悔）。

        Args:
            task_id(number): 事项 id。
        """
        data = await self._api(
            "POST", f"api/v1/tasks/{task_id}/undo", identity=self._identity(event),
        )
        if not data.get("ok", True):
            return f"操作失败：{data.get('body') or data.get('error') or data}"
        return f"已把事项#{task_id} 改回未完成。"

    @filter.llm_tool(name="update_task")
    async def update_task(
        self,
        event: AstrMessageEvent,
        task_id: int,
        content: str = None,
        category: str = None,
        remind_at: str = None,
    ) -> str:
        """修改一条待办的内容/分类/提醒时间。

        Args:
            task_id(number): 事项 id。
            content(string): 新内容（可选）。
            category(string): 新分类（可选）。
            remind_at(string): 新提醒时间（可选）。
        """
        body = {}
        if content:
            body["content"] = content
        if category:
            body["category"] = category
        if remind_at:
            body["remind_at"] = remind_at
        data = await self._api(
            "PATCH", f"api/v1/tasks/{task_id}", identity=self._identity(event), json_body=body,
        )
        if not data.get("ok", True):
            return f"修改失败：{data.get('body') or data.get('error') or data}"
        return f"已修改事项#{task_id}。"

    @filter.llm_tool(name="delete_task")
    async def delete_task(self, event: AstrMessageEvent, task_id: int) -> str:
        """删除一条待办。

        Args:
            task_id(number): 事项 id。
        """
        data = await self._api(
            "DELETE", f"api/v1/tasks/{task_id}", identity=self._identity(event),
        )
        if not data.get("ok", True):
            return f"删除失败：{data.get('body') or data.get('error') or data}"
        return f"已删除事项#{task_id}。"

    @filter.llm_tool(name="export_tasks")
    async def export_tasks(
        self,
        event: AstrMessageEvent,
        is_done: bool = None,
        q: str = None,
        category: str = None,
        from_date: str = None,
        to_date: str = None,
        account_id: int = None,
        destination: str = "text",
        relative_path: str = None,
        limit: int = 20,
    ) -> str:
        """查询待办明细（默认）或导出为 CSV 文件。默认导出全部状态（需过滤请传 is_done）；q 空格分隔多个词须全部命中。纯查询模式返回明细文本供你直接回答，结果过长时只显示前 limit 条并提示导出；destination 可选 workspace（存当前会话工作区）/ user（直接发 CSV 文件）/ both。

        Args:
            is_done(boolean): 状态筛选，true=已完成，false=未完成；不传则导出全部（可选）。
            q(string): 内容关键词，空格分隔多个词须全部命中（可选）。
            category(string): 分类筛选，如 购物（可选）。
            from_date(string): 起始创建日期 YYYY-MM-DD（可选）。
            to_date(string): 结束创建日期 YYYY-MM-DD（可选）。
            account_id(number): 归属账户 id（有多个账本时必须指定）。
            destination(string): text=纯查询返回明细文本（默认）；workspace=只存会话工作区；user=直接把 CSV 文件发给你；both=既存工作区又发文件。
            relative_path(string): destination 为 workspace/both 时，工作区内的相对路径，如 exports/待办清单.csv（可选，默认自动命名）。
            limit(number): 纯查询模式下最多返回的明细条数，默认 20，最大 50（可选）。
        """
        filters = {}
        if is_done is not None:
            filters["is_done"] = "1" if is_done else "0"
        if q:
            filters["q"] = q
        if category:
            filters["category"] = category
        if from_date:
            filters["from"] = from_date
        if to_date:
            filters["to"] = to_date
        if account_id:
            filters["account_id"] = account_id

        # ── 纯查询模式（默认）：返回明细文本给 BOT，不进 LLM 上下文超长风险 ──
        if destination == "text":
            qs = urlencode(filters)
            data = await self._api(
                "GET", f"api/v1/tasks/export?{qs}", identity=self._identity(event)
            )
            if not data.get("ok", True):
                return self._err_msg(data)
            items = data.get("items") or []
            total = data.get("total") or len(items)
            applied = data.get("applied") or {}
            if not items:
                return "没有符合条件的待办。" + (f"\n已生效筛选：{applied}" if applied else "")
            try:
                cap = max(1, min(int(limit) if limit else 20, 50))
            except (TypeError, ValueError):
                cap = 20
            shown = items[:cap]
            lines = [f"共 {total} 条待办（显示前 {len(shown)} 条）："]
            for t in shown:
                done = "✅" if t.get("is_done") else "⬜"
                line = f"- #{t.get('id', '?')} {done} {t.get('content', '')}"
                if t.get("category"):
                    line += f" 分类:{t.get('category')}"
                if t.get("remind_at"):
                    line += f" 提醒:{t.get('remind_at')}"
                lines.append(line)
            if total > len(shown):
                lines.append(
                    f"\n…还有 {total - len(shown)} 条未显示。"
                    f"可用 destination=user 把 CSV 文件发给你，或 destination=workspace 存工作区。"
                )
            if applied:
                lines.append(f"已生效筛选：{applied}")
            return "\n".join(lines)

        # ── 文件导出模式（workspace / user / both）：走 CSV ──
        params = dict(filters)
        params["format"] = "csv"
        qs = urlencode(params)

        ret = await self._api_raw(
            "GET", f"api/v1/tasks/export?{qs}", identity=self._identity(event)
        )
        if isinstance(ret, dict):
            return self._err_msg(ret)
        http_status, csv_text = ret
        if http_status >= 400:
            return f"导出失败（HTTP {http_status}）：{csv_text[:300]}"
        if not isinstance(csv_text, str) or not csv_text.strip():
            return "导出失败：后端返回空内容。"
        try:
            # 用 csv.reader 精确统计行数（正确处理带引号换行的字段），不进 LLM 上下文
            row_count = max(0, sum(1 for _ in csv.reader(io.StringIO(csv_text))) - 1)
        except Exception:
            row_count = 0

        ws_path = None
        if destination in ("workspace", "both"):
            try:
                if relative_path:
                    target = self._resolve_workspace_rel(event, relative_path)
                else:
                    target = (
                        self._workspace_root(event)
                        / "exports"
                        / f"tasks_{datetime.now():%Y%m%d_%H%M%S}.csv"
                    )
                target.parent.mkdir(parents=True, exist_ok=True)
                with open(target, "w", encoding="utf-8", newline="") as f:
                    f.write(csv_text)  # csv_text 自带 UTF-8 BOM
                ws_path = target.relative_to(self._workspace_root(event)).as_posix()
            except ValueError as e:
                return f"无效路径：{e}"

        if destination in ("user", "both"):
            filename = (
                os.path.basename(ws_path)
                if ws_path
                else f"tasks_{datetime.now():%Y%m%d_%H%M%S}.csv"
            )
            tmp = os.path.join(get_astrbot_temp_path(), f"ha_export_{uuid.uuid4().hex}.csv")
            with open(tmp, "w", encoding="utf-8", newline="") as f:
                f.write(csv_text)
            try:
                ok = await self.context.send_message(
                    event.unified_msg_origin,
                    MessageChain([File(name=filename, file=tmp)]),
                )
                if not ok:
                    return f"文件已生成但发送失败（未匹配到平台）。工作区副本：{ws_path or filename}"
            finally:
                if os.path.exists(tmp):
                    os.remove(tmp)

        applied_note = f"筛选：{dict(params)}"
        if destination == "workspace":
            return (
                f"已导出 {row_count} 条待办到工作区：{ws_path}。\n"
                f"{applied_note}\n"
                f"可用文件工具读取/处理，或让我把它发给你。"
            )
        if destination == "user":
            return f"已把 {row_count} 条待办的 CSV 文件发给你。\n{applied_note}"
        return f"已导出 {row_count} 条待办到工作区：{ws_path}，并已把文件发给你。\n{applied_note}"

    # ────────────────────────── 工作账单（装修安装门，2026-08-18） ──────────────────────────

    @staticmethod
    def _fmt_work_bill(ledger: dict) -> str:
        """把后端 work_bill 对账对象（{bill, client_name, items, ...}）渲染成给 LLM 的文本。"""
        b = ledger.get("bill") or {}
        items = ledger.get("items") or []
        item_lines = []
        for it in items:
            it = it or {}
            line = f"- {it.get('name', '')} {it.get('qty', '')}{it.get('unit', '')}×{it.get('unit_price', 0)}分={it.get('amount', 0)}分"
            if it.get("note"):
                line += f"（{it.get('note')}）"
            item_lines.append(line)
        sc = {"settled": "已结算", "partial": "部分结算", "unsettled": "未结算"}.get(ledger.get("status"), ledger.get("status"))
        head = f"账单 #{b.get('id')} {b.get('occurred_at', '')} {ledger.get('client_name', '')}"
        if b.get("address"):
            head += f" {b.get('address')}"
        if b.get("contact"):
            head += f" 联系人:{b.get('contact')}"
        head += f"｜应收 {ledger.get('receivable', 0)}分 已收 {ledger.get('paid', 0)}分 欠 {ledger.get('owed', 0)}分（{sc}）"
        if b.get("final_amount") is not None:
            head += f"｜实际应收 {b.get('final_amount')}分"
        return head + ("\n" + "\n".join(item_lines) if item_lines else "")

    @filter.llm_tool(name="add_work_client")
    async def add_work_client(
        self,
        event: AstrMessageEvent,
        name: str,
        type: str = "company",
        phone: str = None,
        note: str = None,
        account_id: int = None,
    ) -> str:
        """新建一个委托方（为谁做事/跟谁结算：装修公司或个人）。

        Args:
            name(string): 委托方名称，如 A装修公司 / 张三。
            type(string): company=装修公司（默认）；personal=个人委托。
            phone(string): 联系电话（可选）。
            note(string): 备注（可选）。
            account_id(number): 归属账户 id（有多个账本时必须指定）。
        """
        body = {"name": name, "type": type}
        if phone is not None:
            body["phone"] = phone
        if note is not None:
            body["note"] = note
        qs = f"?account_id={account_id}" if account_id else ""
        data = await self._api("POST", f"api/v1/work-clients{qs}", identity=self._identity(event), json_body=body)
        if not data.get("ok", True):
            return self._err_msg(data)
        c = data.get("client") or {}
        return f"已建委托方 #{c.get('id')}：{c.get('name')}（{'装修公司' if c.get('type') == 'company' else '个人委托'}）"

    @filter.llm_tool(name="list_work_clients")
    async def list_work_clients(self, event: AstrMessageEvent, q: str = None, account_id: int = None) -> str:
        """列出委托方。

        Args:
            q(string): 名称/电话关键词（可选）。
            account_id(number): 归属账户 id（有多个账本时必须指定）。
        """
        params = {}
        if q:
            params["q"] = q
        if account_id:
            params["account_id"] = account_id
        qs = urlencode(params)
        data = await self._api("GET", f"api/v1/work-clients?{qs}", identity=self._identity(event))
        if not data.get("ok", True):
            return self._err_msg(data)
        items = data.get("items") or []
        if not items:
            return "暂无委托方。"
        lines = []
        for c in items[:20]:
            kind = "装修公司" if c.get("type") == "company" else "个人委托"
            phone = f" 电话:{c.get('phone')}" if c.get("phone") else ""
            lines.append(f"#{c.get('id')} {c.get('name')}（{kind}）{phone}")
        return "委托方：\n" + "\n".join(lines)

    @filter.llm_tool(name="update_work_client")
    async def update_work_client(
        self,
        event: AstrMessageEvent,
        client_id: int,
        name: str = None,
        type: str = None,
        phone: str = None,
        note: str = None,
    ) -> str:
        """修改委托方信息。

        Args:
            client_id(number): 委托方 id。
            name(string): 新名称（可选）。
            type(string): company 或 personal（可选）。
            phone(string): 新电话（可选）。
            note(string): 新备注（可选）。
        """
        body = {}
        if name is not None:
            body["name"] = name
        if type is not None:
            body["type"] = type
        if phone is not None:
            body["phone"] = phone
        if note is not None:
            body["note"] = note
        data = await self._api("PATCH", f"api/v1/work-clients/{client_id}", identity=self._identity(event), json_body=body)
        if not data.get("ok", True):
            return self._err_msg(data)
        c = data.get("client") or {}
        return f"已更新委托方 #{c.get('id')}：{c.get('name')}"

    @filter.llm_tool(name="delete_work_client")
    async def delete_work_client(self, event: AstrMessageEvent, client_id: int) -> str:
        """删除一个委托方（其下还有未删除账单时不可删）。

        Args:
            client_id(number): 委托方 id。
        """
        data = await self._api("DELETE", f"api/v1/work-clients/{client_id}", identity=self._identity(event))
        if not data.get("ok", True):
            return self._err_msg(data)
        return f"已删除委托方 #{client_id}。"

    @filter.llm_tool(name="set_work_price")
    async def set_work_price(
        self,
        event: AstrMessageEvent,
        client_id: int,
        name: str,
        unit_price: int,
        unit: str = None,
        note: str = None,
        account_id: int = None,
    ) -> str:
        """设置某委托方的安装内容单价（已有同名则更新）。价格按委托方区分。

        Args:
            client_id(number): 委托方 id。
            name(string): 品名，如 欧派木门 / 双包卫生间门 / 进户门套 / 隐形门 / 墙板。
            unit_price(number): 单价（单位：分，1 元 = 100 分）。
            unit(string): 单位，如 扇/个/套/㎡（可选，默认 个）。
            note(string): 备注（可选）。
            account_id(number): 归属账户 id（有多个账本时必须指定）。
        """
        body = {"client_id": client_id, "name": name, "unit_price": unit_price}
        if unit is not None:
            body["unit"] = unit
        if note is not None:
            body["note"] = note
        qs = f"?account_id={account_id}" if account_id else ""
        data = await self._api("POST", f"api/v1/work-unit-prices{qs}", identity=self._identity(event), json_body=body)
        if not data.get("ok", True):
            return self._err_msg(data)
        p = data.get("price") or {}
        return f"已设置单价：{p.get('name')} {p.get('unit_price')}分/{p.get('unit')}（委托方#{p.get('client_id')}）"

    @filter.llm_tool(name="list_work_prices")
    async def list_work_prices(
        self, event: AstrMessageEvent, client_id: int = None, q: str = None, account_id: int = None
    ) -> str:
        """列出单价表。

        Args:
            client_id(number): 按委托方过滤（可选）。
            q(string): 品名关键词（可选）。
            account_id(number): 归属账户 id（有多个账本时必须指定）。
        """
        params = {}
        if client_id:
            params["client_id"] = client_id
        if q:
            params["q"] = q
        if account_id:
            params["account_id"] = account_id
        qs = urlencode(params)
        data = await self._api("GET", f"api/v1/work-unit-prices?{qs}", identity=self._identity(event))
        if not data.get("ok", True):
            return self._err_msg(data)
        items = data.get("items") or []
        if not items:
            return "暂无单价记录。"
        lines = [f"#{p.get('id')} {p.get('name')} {p.get('unit_price')}分/{p.get('unit')}（委托方#{p.get('client_id')}）" for p in items[:50]]
        return "单价表：\n" + "\n".join(lines)

    @filter.llm_tool(name="delete_work_price")
    async def delete_work_price(self, event: AstrMessageEvent, price_id: int, account_id: int = None) -> str:
        """删除一条单价记录（历史账单单价快照不受影响，重算时跳过）。

        Args:
            price_id(number): 单价记录 id。
            account_id(number): 归属账户 id（有多个账本时必须指定）。
        """
        qs = f"?account_id={account_id}" if account_id else ""
        data = await self._api("DELETE", f"api/v1/work-unit-prices/{price_id}{qs}", identity=self._identity(event))
        if not data.get("ok", True):
            return self._err_msg(data)
        return f"已删除单价 #{price_id}。"

    @filter.llm_tool(name="add_work_bill")
    async def add_work_bill(
        self,
        event: AstrMessageEvent,
        client_id: int,
        items: list,
        address: str = None,
        contact: str = None,
        occurred_at: str = None,
        note: str = None,
        final_amount: int = None,
        account_id: int = None,
    ) -> str:
        """新建一张工作账单（安装门的工作记录）。金额 = Σ明细小计；明细单价缺省自动按该委托方单价表带出，可覆盖。

        Args:
            client_id(number): 委托方 id。
            items(list): 安装内容明细行数组，每项 {name: 品名, qty?: 数量(可小数,默认1), unit?: 单位, unit_price?: 单价(分,缺省从单价表带出), note?: 行备注}。
            address(string): 安装地点（可选）。
            contact(string): 服务对象/联系人姓名（可选）。
            occurred_at(string): 安装日期 YYYY-MM-DD（可选，默认今天）。
            note(string): 整单备注（可选）。
            final_amount(number): 实际应收（分，可选；缺省按明细计算）。
            account_id(number): 归属账户 id（有多个账本时必须指定）。
        """
        body = {"client_id": client_id, "items": items}
        if address is not None:
            body["address"] = address
        if contact is not None:
            body["contact"] = contact
        if occurred_at is not None:
            body["occurred_at"] = occurred_at
        if note is not None:
            body["note"] = note
        if final_amount is not None:
            body["final_amount"] = final_amount
        qs = f"?account_id={account_id}" if account_id else ""
        data = await self._api("POST", f"api/v1/work-bills{qs}", identity=self._identity(event), json_body=body)
        if not data.get("ok", True):
            return self._err_msg(data)
        return self._fmt_work_bill(data.get("bill") or {})

    @filter.llm_tool(name="list_work_bills")
    async def list_work_bills(
        self,
        event: AstrMessageEvent,
        client_id: int = None,
        contact: str = None,
        keyword: str = None,
        status: str = None,
        from_date: str = None,
        to_date: str = None,
        account_id: int = None,
        page: int = None,
        page_size: int = None,
    ) -> str:
        """列出工作账单，支持按委托方/联系人/地址备注/状态/安装日期过滤 + 分页，回显生效筛选。

        Args:
            client_id(number): 委托方 id（可选）。
            contact(string): 服务对象/联系人姓名关键词（可选）。
            keyword(string): 地址或备注包含的关键词，空格分隔多个词须全部命中（可选）。
            status(string): 状态，unsettled=未结算 / partial=部分结算 / settled=已结算（可选）。
            from_date(string): 起始安装日期 YYYY-MM-DD（可选）。
            to_date(string): 结束安装日期 YYYY-MM-DD（可选）。
            account_id(number): 归属账户 id（有多个账本时必须指定）。
            page(number): 页码，从 1 开始（可选）。
            page_size(number): 每页条数，默认 50，最大 200（可选）。
        """
        params = {}
        if client_id:
            params["client_id"] = client_id
        if contact:
            params["contact"] = contact
        if keyword:
            params["keyword"] = keyword
        if status:
            params["status"] = status
        if from_date:
            params["from"] = from_date
        if to_date:
            params["to"] = to_date
        if account_id:
            params["account_id"] = account_id
        if page:
            params["page"] = page
        if page_size:
            params["page_size"] = page_size
        qs = urlencode(params)
        data = await self._api("GET", f"api/v1/work-bills?{qs}", identity=self._identity(event))
        if not data.get("ok", True):
            return self._err_msg(data)
        items = data.get("items") or []
        total = data.get("total") or len(items)
        applied = data.get("applied") or {}
        if not items:
            return "没有符合条件的账单。" + (f"\n已生效筛选：{applied}" if applied else "")
        lines = [f"共 {total} 张工作账单（显示前 {len(items)} 条）："]
        for it in items[:20]:
            sc = {"settled": "已结", "partial": "部分结", "unsettled": "未结"}.get(it.get("status"), it.get("status"))
            lines.append(
                f"#{it.get('id')} {it.get('occurred_at', '')} {it.get('client_name', '')} "
                f"{it.get('address', '')} {it.get('contact', '')}｜应收{it.get('receivable', 0)}分 已收{it.get('paid', 0)}分 {sc}"
            )
        if total > len(items):
            lines.append(f"\n…还有 {total - len(items)} 张未显示，可缩小条件或用 export_work_bills 导出。")
        if applied:
            lines.append(f"已生效筛选：{applied}")
        return "\n".join(lines)

    @filter.llm_tool(name="get_work_bill")
    async def get_work_bill(self, event: AstrMessageEvent, bill_id: int) -> str:
        """查看一张工作账单的完整明细（含安装内容式子、应收/已收/欠款/状态）。

        Args:
            bill_id(number): 工作账单 id。
        """
        data = await self._api("GET", f"api/v1/work-bills/{bill_id}", identity=self._identity(event))
        if not data.get("ok", True):
            return self._err_msg(data)
        return self._fmt_work_bill(data.get("bill") or {})

    @filter.llm_tool(name="update_work_bill")
    async def update_work_bill(
        self,
        event: AstrMessageEvent,
        bill_id: int,
        client_id: int = None,
        address: str = None,
        contact: str = None,
        occurred_at: str = None,
        note: str = None,
        final_amount: int = None,
        items: list = None,
    ) -> str:
        """修改工作账单（传 items 时全量替换明细，务必传全量）。

        Args:
            bill_id(number): 工作账单 id。
            client_id(number): 改委托方（可选）。
            address(string): 安装地点（可选）。
            contact(string): 服务对象（可选）。
            occurred_at(string): 安装日期 YYYY-MM-DD（可选）。
            note(string): 整单备注（可选）。
            final_amount(number): 实际应收（分，可选；传 0 清除回按明细计算）。
            items(list): 全量替换明细（格式同 add_work_bill 的 items，传则整组替换）。
        """
        body = {}
        if client_id is not None:
            body["client_id"] = client_id
        if address is not None:
            body["address"] = address
        if contact is not None:
            body["contact"] = contact
        if occurred_at is not None:
            body["occurred_at"] = occurred_at
        if note is not None:
            body["note"] = note
        if final_amount is not None:
            body["final_amount"] = final_amount if final_amount != 0 else None  # 0 = 清除
        if items is not None:
            body["items"] = items
        data = await self._api("PATCH", f"api/v1/work-bills/{bill_id}", identity=self._identity(event), json_body=body)
        if not data.get("ok", True):
            return self._err_msg(data)
        return self._fmt_work_bill(data.get("bill") or {})

    @filter.llm_tool(name="delete_work_bill")
    async def delete_work_bill(self, event: AstrMessageEvent, bill_id: int) -> str:
        """删除一张工作账单（软删）。

        Args:
            bill_id(number): 工作账单 id。
        """
        data = await self._api("DELETE", f"api/v1/work-bills/{bill_id}", identity=self._identity(event))
        if not data.get("ok", True):
            return self._err_msg(data)
        return f"已删除工作账单 #{bill_id}。"

    @filter.llm_tool(name="settle_work_bill")
    async def settle_work_bill(
        self,
        event: AstrMessageEvent,
        bill_id: int,
        amount: int,
        settled_at: str = None,
        note: str = None,
    ) -> str:
        """记录一笔结算实收（可多次、可部分；金额记实际收到，与计算金额解耦）。

        Args:
            bill_id(number): 工作账单 id。
            amount(number): 本次实收金额（单位：分，1 元 = 100 分）。
            settled_at(string): 结算日期 YYYY-MM-DD（可选，默认今天）。
            note(string): 备注（可选）。
        """
        body = {"amount": amount}
        if settled_at is not None:
            body["settled_at"] = settled_at
        if note is not None:
            body["note"] = note
        data = await self._api("POST", f"api/v1/work-bills/{bill_id}/settle", identity=self._identity(event), json_body=body)
        if not data.get("ok", True):
            return self._err_msg(data)
        return f"已记录结算 {amount} 分。\n" + self._fmt_work_bill(data.get("bill") or {})

    @filter.llm_tool(name="recalc_work_bills")
    async def recalc_work_bills(
        self,
        event: AstrMessageEvent,
        bill_ids: list = None,
        client_id: int = None,
        from_date: str = None,
        to_date: str = None,
        apply: bool = False,
    ) -> str:
        """按最新单价表批量重算未结算账单（已结算单锁定不动）。改单价后账单不会自动变，需要时手动触发；默认只预览返回 diff，apply=true 才提交。

        Args:
            bill_ids(list): 只重算指定的账单 id 数组（可选；不给则按 client_id/日期筛全部未结算单）。
            client_id(number): 按委托方筛全部未结算单（可选）。
            from_date(string): 起始安装日期 YYYY-MM-DD（可选）。
            to_date(string): 结束安装日期 YYYY-MM-DD（可选）。
            apply(boolean): true=提交重算；默认 false=只预览（可选）。
        """
        params = {}
        if bill_ids:
            params["bill_ids"] = ",".join(str(x) for x in bill_ids)
        if client_id:
            params["client_id"] = client_id
        if from_date:
            params["from"] = from_date
        if to_date:
            params["to"] = to_date
        if apply:
            params["dry_run"] = "0"
        qs = urlencode(params)
        data = await self._api("POST", f"api/v1/work-bills/recalc?{qs}", identity=self._identity(event))
        if not data.get("ok", True):
            return self._err_msg(data)
        changed = data.get("changed") or []
        dry = data.get("dry_run", True)
        if not changed:
            return "重算预览：没有明细需要更新（无未结算单或单价未变）。"
        lines = [f"重算{'预览' if dry else '完成'}：{data.get('affected_bills', 0)} 张账单共 {data.get('total_changes', 0)} 处明细变更："]
        for c in changed[:30]:
            lines.append(
                f"- 账单#{c.get('bill_id')} {c.get('name')} {c.get('qty')}{c.get('unit')}："
                f"{c.get('old_price')}分→{c.get('new_price')}分（{c.get('old_amount')}分→{c.get('new_amount')}分）"
            )
        if dry:
            lines.append("\n这是预览，未实际修改。确认后可用 recalc_work_bills(apply=true) 提交。")
        if data.get("applied"):
            lines.append(f"已生效筛选：{data.get('applied')}")
        return "\n".join(lines)

    @filter.llm_tool(name="export_work_bills")
    async def export_work_bills(
        self,
        event: AstrMessageEvent,
        mode: str = "summary",
        client_id: int = None,
        keyword: str = None,
        from_date: str = None,
        to_date: str = None,
        status: str = None,
        account_id: int = None,
        destination: str = "text",
        relative_path: str = None,
        limit: int = 20,
    ) -> str:
        """导出工作账单。summary=日常简版（明细平铺）；statement=结账版（含计算式子+对账）。纯查询返回文本；也可存工作区/发 CSV 文件。

        Args:
            mode(string): summary=日常简版（默认）；statement=结账版（带式子）。
            client_id(number): 按委托方过滤（可选）。
            keyword(string): 地址或备注包含的关键词（可选）。
            from_date(string): 起始安装日期 YYYY-MM-DD（可选）。
            to_date(string): 结束安装日期 YYYY-MM-DD（可选）。
            status(string): unsettled/partial/settled（可选）。
            account_id(number): 归属账户 id（有多个账本时必须指定）。
            destination(string): text=纯查询返回文本（默认）；workspace=只存会话工作区；user=直接把 CSV 文件发给你；both=两者。
            relative_path(string): destination 为 workspace/both 时的工作区内相对路径（可选，默认自动命名）。
            limit(number): 纯查询模式下最多返回账单条数，默认 20，最大 50（可选）。
        """
        filters = {"mode": mode}
        if client_id:
            filters["client_id"] = client_id
        if keyword:
            filters["keyword"] = keyword
        if from_date:
            filters["from"] = from_date
        if to_date:
            filters["to"] = to_date
        if status:
            filters["status"] = status
        if account_id:
            filters["account_id"] = account_id

        if destination == "text":
            qs = urlencode(filters)
            data = await self._api("GET", f"api/v1/work-bills/export?{qs}", identity=self._identity(event))
            if not data.get("ok", True):
                return self._err_msg(data)
            bills = data.get("bills") or []
            total = data.get("total") or len(bills)
            applied = data.get("applied") or {}
            if not bills:
                return "没有符合条件的账单。" + (f"\n已生效筛选：{applied}" if applied else "")
            try:
                cap = max(1, min(int(limit) if limit else 20, 50))
            except (TypeError, ValueError):
                cap = 20
            shown = bills[:cap]
            lines = [f"共 {total} 张账单（显示前 {len(shown)} 张）："]
            for b in shown:
                sc = {"settled": "已结", "partial": "部分结", "unsettled": "未结"}.get(b.get("status"), b.get("status"))
                if mode == "statement":
                    lines.append(f"#{b.get('id')} {b.get('occurred_at', '')} {b.get('client_name', '')} {b.get('address', '')} {b.get('contact', '')}")
                    lines.append(f"  式子：{b.get('formula', '')}")
                    lines.append(f"  计算{b.get('computed_total', 0)}分｜应收{b.get('receivable', 0)}分｜已收{b.get('paid', 0)}分｜欠{b.get('owed', 0)}分（{sc}）")
                else:
                    for it in b.get("items") or []:
                        lines.append(f"#{b.get('id')} {b.get('occurred_at', '')} {b.get('client_name', '')}｜{it.get('name')} {it.get('qty')}{it.get('unit')} {it.get('amount', 0)}分")
            if total > len(shown):
                lines.append(f"\n…还有 {total - len(shown)} 张未显示，可用 destination=user 发文件或 destination=workspace 存工作区。")
            if applied:
                lines.append(f"已生效筛选：{applied}")
            return "\n".join(lines)

        # 文件导出模式：CSV
        params = dict(filters)
        params["format"] = "csv"
        qs = urlencode(params)
        ret = await self._api_raw("GET", f"api/v1/work-bills/export?{qs}", identity=self._identity(event))
        if isinstance(ret, dict):
            return self._err_msg(ret)
        http_status, csv_text = ret
        if http_status >= 400:
            return f"导出失败（HTTP {http_status}）：{csv_text[:300]}"
        if not isinstance(csv_text, str) or not csv_text.strip():
            return "导出失败：后端返回空内容。"
        try:
            row_count = max(0, sum(1 for _ in csv.reader(io.StringIO(csv_text))) - 1)
        except Exception:
            row_count = 0

        ws_path = None
        if destination in ("workspace", "both"):
            try:
                if relative_path:
                    target = self._resolve_workspace_rel(event, relative_path)
                else:
                    target = (
                        self._workspace_root(event)
                        / "exports"
                        / f"work-bills-{mode}_{datetime.now():%Y%m%d_%H%M%S}.csv"
                    )
                target.parent.mkdir(parents=True, exist_ok=True)
                with open(target, "w", encoding="utf-8", newline="") as f:
                    f.write(csv_text)
                ws_path = target.relative_to(self._workspace_root(event)).as_posix()
            except ValueError as e:
                return f"无效路径：{e}"

        if destination in ("user", "both"):
            filename = (
                os.path.basename(ws_path)
                if ws_path
                else f"work-bills-{mode}_{datetime.now():%Y%m%d_%H%M%S}.csv"
            )
            tmp = os.path.join(get_astrbot_temp_path(), f"ha_export_{uuid.uuid4().hex}.csv")
            with open(tmp, "w", encoding="utf-8", newline="") as f:
                f.write(csv_text)
            try:
                ok = await self.context.send_message(
                    event.unified_msg_origin,
                    MessageChain([File(name=filename, file=tmp)]),
                )
                if not ok:
                    return f"文件已生成但发送失败（未匹配到平台）。工作区副本：{ws_path or filename}"
            finally:
                if os.path.exists(tmp):
                    os.remove(tmp)

        applied_note = f"筛选：{dict(params)}"
        if destination == "workspace":
            return (
                f"已导出 {row_count} 行到工作区：{ws_path}。\n"
                f"{applied_note}\n"
                f"可用文件工具读取/处理，或让我把它发给你。"
            )
        if destination == "user":
            return f"已把 {row_count} 行的 CSV 文件发给你。\n{applied_note}"
        return f"已导出 {row_count} 行到工作区：{ws_path}，并已把文件发给你。\n{applied_note}"

    @filter.llm_tool(name="work_bill_stats")
    async def work_bill_stats(
        self,
        event: AstrMessageEvent,
        year: int = None,
        month: int = None,
        from_date: str = None,
        to_date: str = None,
        client_id: int = None,
        status: str = None,
        account_id: int = None,
        chart: bool = False,
    ) -> str:
        """工作账单统计：合计应收/已收/欠款 + 按委托方 + 按月份；chart=True 附统计图。

        Args:
            year(number): 年份 YYYY（可选）。
            month(number): 月份 1-12（可选，需配 year）。
            from_date(string): 起始安装日期 YYYY-MM-DD（可选，优先于 year/month）。
            to_date(string): 结束安装日期 YYYY-MM-DD（可选）。
            client_id(number): 按委托方过滤（可选）。
            status(string): unsettled/partial/settled（可选）。
            account_id(number): 归属账户 id（有多个账本时必须指定）。
            chart(boolean): 是否发送统计图 PNG（默认 false）。
        """
        params = {}
        if year is not None:
            params["year"] = year
        if month is not None:
            params["month"] = month
        if from_date:
            params["from"] = from_date
        if to_date:
            params["to"] = to_date
        if client_id:
            params["client_id"] = client_id
        if status:
            params["status"] = status
        if account_id:
            params["account_id"] = account_id
        qs = urlencode(params)
        data = await self._api("GET", f"api/v1/work-bills/stats?{qs}", identity=self._identity(event))
        if not data.get("ok", True):
            return self._err_msg(data)

        def _y(v):
            return (v or 0) / 100

        lines = [
            f"工作账单统计：共 {data.get('bill_count', 0)} 张｜"
            f"应收 {_y(data.get('receivable')):,.2f}元｜已收 {_y(data.get('paid')):,.2f}元｜欠款 {_y(data.get('owed')):,.2f}元"
        ]
        by_client = data.get("by_client") or []
        if by_client:
            lines.append("按委托方：")
            for c in by_client[:8]:
                lines.append(
                    f"  {c.get('client_name')} {c.get('bill_count')}张｜应收{_y(c.get('receivable')):,.2f} "
                    f"已收{_y(c.get('paid')):,.2f} 欠{_y(c.get('owed')):,.2f}元"
                )
        by_month = data.get("by_month") or []
        if by_month:
            lines.append("按月份：")
            for m in by_month[:12]:
                lines.append(f"  {m.get('month')} {m.get('bill_count')}张｜应收{_y(m.get('receivable')):,.2f} 已收{_y(m.get('paid')):,.2f}元")
        if data.get("applied"):
            lines.append(f"已生效筛选：{data.get('applied')}")
        text = "\n".join(lines)

        if chart and (data.get("receivable") or data.get("paid")):
            if not self._ensure_matplotlib():
                return text + "\n（图表生成失败，仅返回文字）"
            png = os.path.join(get_astrbot_temp_path(), f"ha_work_stats_{uuid.uuid4().hex}.png")
            try:
                await asyncio.to_thread(self._render_work_stats_png, data, str(data.get("applied") or {}), png)
                caption = f"工作账单统计（单位：元）· 应收 {_y(data.get('receivable')):,.2f} 元"
                await self.context.send_message(
                    event.unified_msg_origin,
                    MessageChain().message(caption).file_image(png),
                )
            except Exception as e:
                logger.error(f"[HomeAssistant] 发送工作统计图失败: {e}")
            finally:
                if os.path.exists(png):
                    os.remove(png)
        return text

    # ────────────────────────── 订阅与新闻 ──────────────────────────

    @filter.llm_tool(name="subscribe_source")
    async def subscribe_source(
        self,
        event: AstrMessageEvent,
        source_type: str,
        name: str = None,
        source_url: str = None,
        preset_key: str = None,
    ) -> str:
        """订阅一个信息源（RSS 或预设适配器）。

        Args:
            source_type(string): rss 或 preset。
            name(string): 订阅名称（可选）。
            source_url(string): RSS 地址（source_type=rss 时必填）。
            preset_key(string): 预设适配器键名，如 steam_news（source_type=preset 时必填）。
        """
        body = {"source_type": source_type}
        if name:
            body["name"] = name
        if source_url:
            body["source_url"] = source_url
        if preset_key:
            body["preset_key"] = preset_key
        data = await self._api(
            "POST", "api/v1/subscriptions", identity=self._identity(event), json_body=body,
        )
        if not data.get("ok", True):
            return f"订阅失败：{data.get('body') or data.get('error') or data}"
        s = data.get("subscription") or {}
        return f"已订阅：{s.get('name') or name}（#{s.get('id')}）"

    @filter.llm_tool(name="list_subscriptions")
    async def list_subscriptions(self, event: AstrMessageEvent) -> str:
        """列出当前用户的订阅源。"""
        data = await self._api(
            "GET", "api/v1/subscriptions", identity=self._identity(event),
        )
        if not data.get("ok", True):
            return f"查询失败：{data.get('body') or data.get('error') or data}"
        subs = data.get("subscriptions") or []
        if not subs:
            return "暂无订阅。"
        lines = [
            f"#{s.get('id')} {s.get('name')}（{s.get('source_type')}）"
            for s in subs
        ]
        return "订阅源：\n" + "\n".join(lines)

    @filter.llm_tool(name="unsubscribe")
    async def unsubscribe(self, event: AstrMessageEvent, subscription_id: int) -> str:
        """退订一个信息源。

        Args:
            subscription_id(number): 订阅 id。
        """
        data = await self._api(
            "DELETE", f"api/v1/subscriptions/{subscription_id}", identity=self._identity(event),
        )
        if not data.get("ok", True):
            return f"退订失败：{data.get('body') or data.get('error') or data}"
        return f"已退订订阅#{subscription_id}。"

    @filter.llm_tool(name="query_news")
    async def query_news(
        self,
        event: AstrMessageEvent,
        subscription_id: int = None,
        limit: int = None,
    ) -> str:
        """查询已缓存的信息源新闻。

        Args:
            subscription_id(number): 指定订阅的新闻（可选）。
            limit(number): 返回条数（可选，默认 10）。
        """
        params = {}
        if subscription_id:
            params["subscription_id"] = subscription_id
        if limit:
            params["limit"] = limit
        qs = "&".join(f"{k}={v}" for k, v in params.items())
        data = await self._api(
            "GET", f"api/v1/news?{qs}", identity=self._identity(event),
        )
        if not data.get("ok", True):
            return f"查询失败：{data.get('body') or data.get('error') or data}"
        items = data.get("items") or []
        if not items:
            return "暂无新闻。"
        lines = []
        for it in items[:limit or 10]:
            lines.append(f"- {it.get('title') or it.get('summary') or '(无标题)'}")
        return "新闻：\n" + "\n".join(lines)

    # ────────────────────────── 实时抓取 ──────────────────────────

    @filter.llm_tool(name="fetch_source")
    async def fetch_source(
        self,
        event: AstrMessageEvent,
        source_type: str,
        source_url: str = None,
        preset_key: str = None,
        params: dict = None,
    ) -> str:
        """实时抓取一次信息源（RSS/URL/预设适配器）。

        Args:
            source_type(string): rss、url 或 preset。
            source_url(string): 目标 URL（source_type=rss|url 时用）。
            preset_key(string): 预设适配器键名（source_type=preset 时用）。
            params(object): 附加参数（可选）。
        """
        body = {"source_type": source_type}
        if source_url:
            body["source_url"] = source_url
        if preset_key:
            body["preset_key"] = preset_key
        if params:
            body["params"] = params
        data = await self._api(
            "POST", "api/v1/fetch", identity=self._identity(event), json_body=body,
        )
        if not data.get("ok", True):
            return f"抓取失败：{data.get('body') or data.get('error') or data}"
        items = data.get("items") or []
        lines = []
        for it in items[:10]:
            lines.append(f"- {it.get('title') or it.get('summary') or '(无标题)'}")
        return f"抓取到 {data.get('count', len(items))} 条：\n" + "\n".join(lines)

    # ────────────────────────── 身份/账户补充 ──────────────────────────

    @filter.llm_tool(name="bind_person")
    async def bind_person(self, event: AstrMessageEvent, person_id: int) -> str:
        """把当前用户绑定到已有的账号（如家人把你拉进家庭）。

        Args:
            person_id(number): 要绑定的账号 id。
        """
        data = await self._api(
            "POST", f"api/v1/persons/{person_id}/bind", identity=self._identity(event),
        )
        if not data.get("ok", True):
            return f"绑定失败：{data.get('body') or data.get('error') or data}"
        return f"绑定成功。"

    @filter.llm_tool(name="join_account")
    async def join_account(self, event: AstrMessageEvent, account_id: int) -> str:
        """加入一个家庭账本。

        Args:
            account_id(number): 家庭账本 id。
        """
        data = await self._api(
            "POST", f"api/v1/accounts/{account_id}/join", identity=self._identity(event),
        )
        if not data.get("ok", True):
            return f"加入失败：{data.get('body') or data.get('error') or data}"
        return f"已加入账本#{account_id}。"

    @filter.llm_tool(name="unbind_identity")
    async def unbind_identity(
        self,
        event: AstrMessageEvent,
        person_id: int,
        platform: str,
        openid: str,
    ) -> str:
        """把某平台身份从账号解绑。

        Args:
            person_id(number): 账号 id。
            platform(string): 平台，qq 或 wechat。
            openid(string): 要解绑的 openid。
        """
        data = await self._api(
            "DELETE", f"api/v1/persons/{person_id}/identities/{platform}/{openid}",
            identity=self._identity(event),
        )
        if not data.get("ok", True):
            return f"解绑失败：{data.get('body') or data.get('error') or data}"
        return f"已解绑 {platform}/{openid}。"

    @filter.llm_tool(name="set_primary_identity")
    async def set_primary_identity(
        self,
        event: AstrMessageEvent,
        platform: str,
        openid: str,
    ) -> str:
        """把主平台切换为本人名下另一身份。

        Args:
            platform(string): 平台，qq 或 wechat。
            openid(string): 目标身份 openid。
        """
        data = await self._api(
            "PATCH", "api/v1/persons/me/primary-identity",
            identity=self._identity(event), json_body={"platform": platform, "openid": openid},
        )
        if not data.get("ok", True):
            return f"切换失败：{data.get('body') or data.get('error') or data}"
        return f"主身份已切换为 {platform}/{openid}。"

    # ────────────────────────── 记账 Web 链接 ──────────────────────────

    @filter.llm_tool(name="create_ledger_link")
    async def create_ledger_link(
        self,
        event: AstrMessageEvent,
        mode: str = "read",
        expires_in: int = None,
    ) -> str:
        """生成一个可查看/修改账本的网页链接（家人可在浏览器打开）。

        Args:
            mode(string): read 或 write，默认 read。
            expires_in(number): 有效期（分钟，1-1440），默认 30。
        """
        body = {"mode": mode}
        if expires_in:
            body["expires_in"] = int(expires_in)
        data = await self._api(
            "POST", "api/v1/web/tokens", identity=self._identity(event), json_body=body,
        )
        if not data.get("ok", True):
            return f"生成失败：{data.get('body') or data.get('error') or data}"
        return (
            f"账本链接（{data.get('mode')}，{data.get('expires_at')} 过期）：\n{data.get('url')}"
        )

    @filter.llm_tool(name="revoke_ledger_link")
    async def revoke_ledger_link(self, event: AstrMessageEvent, token_id: int) -> str:
        """撤销一个已生成的账本链接。

        Args:
            token_id(number): 链接 token 的 id。
        """
        data = await self._api(
            "DELETE", f"api/v1/web/tokens/{token_id}", identity=self._identity(event),
        )
        if not data.get("ok", True):
            return f"撤销失败：{data.get('body') or data.get('error') or data}"
        return f"已撤销链接#{token_id}。"

    @filter.llm_tool(name="list_ledger_links")
    async def list_ledger_links(self, event: AstrMessageEvent) -> str:
        """列出当前用户已生成的账本网页链接（含过期状态）。

        Args:
        """
        data = await self._api(
            "GET", "api/v1/web/tokens", identity=self._identity(event),
        )
        if not data.get("ok", True):
            return f"查询失败：{data.get('body') or data.get('error') or data}"
        tokens = data.get("tokens") or []
        if not tokens:
            return "还没有生成过账本链接。"
        lines = []
        for t in tokens[:10]:
            state = "已过期" if t.get("is_expired") else "有效"
            lines.append(
                f"#{t.get('id')} {t.get('mode')} {state} 过期:{t.get('expires_at')}"
            )
        return "账本链接：\n" + "\n".join(lines)

    # ────────────────────────── 统计 / 导出（方向 A） ──────────────────────────

    @filter.llm_tool(name="bill_stats")
    async def bill_stats(
        self,
        event: AstrMessageEvent,
        year: int = None,
        month: int = None,
        from_date: str = None,
        to_date: str = None,
        category: str = None,
        amount_min: int = None,
        amount_max: int = None,
        account_id: int = None,
        chart: bool = True,
    ) -> str:
        """查询收支统计（可含图表）。按年份/月份/日期区间/分类/金额区间筛选；chart=True 时把统计图（支出构成+收支趋势）直接发给你。

        Args:
            year(number): 年份，如 2026（可选）。
            month(number): 月份 1-12，需配合 year（可选）。
            from_date(string): 起始日期 YYYY-MM-DD，优先于 year/month（可选）。
            to_date(string): 结束日期 YYYY-MM-DD，与 from_date 成对（可选）。
            category(string): 只看某一分类，如 餐饮（可选）。
            amount_min(number): 金额下限（单位：分，1 元 = 100 分）（可选）。
            amount_max(number): 金额上限（单位：分，1 元 = 100 分）（可选）。
            account_id(number): 记账账户 id（有多个账本时必须指定，否则返回 409）。
            chart(boolean): 是否同时生成并发送统计图，默认 true。
        """
        params = {}
        if year:
            params["year"] = year
        if month:
            params["month"] = month
        if from_date:
            params["from"] = from_date
        if to_date:
            params["to"] = to_date
        if category:
            params["category"] = category
        if amount_min is not None:
            params["amount_min"] = amount_min
        if amount_max is not None:
            params["amount_max"] = amount_max
        if account_id:
            params["account_id"] = account_id
        qs = urlencode(params)
        data = await self._api(
            "GET", f"api/v1/bills/stats/range?{qs}", identity=self._identity(event)
        )
        if not data.get("ok", True):
            return self._err_msg(data)
        income = int(data.get("income") or 0)
        expense = int(data.get("expense") or 0)
        net = int(data.get("net") or 0)
        by_cat = data.get("by_category") or []
        trend = data.get("trend") or []
        applied = data.get("applied") or {}

        if income == 0 and expense == 0:
            return "该筛选区间暂无账单。"

        lines = ["收支统计："]
        lines.append(f"收入 {income} 分 / 支出 {expense} 分 / 结余 {net} 分")
        top = sorted(
            [c for c in by_cat if c.get("amount")],
            key=lambda c: c["amount"],
            reverse=True,
        )[:5]
        if top:
            lines.append(
                "支出分类 Top：" + "、".join(
                    f"{c['category']} {c['amount']}分({c.get('count', 0)}笔)"
                    for c in top
                )
            )
            if len(by_cat) > 5:
                lines.append(f"…共 {len(by_cat)} 个分类")
        if trend:
            lines.append(
                "趋势：" + "；".join(
                    f"{t.get('month', '')} 收{t.get('income', 0)}/支{t.get('expense', 0)}"
                    for t in trend
                )
            )
        if applied:
            lines.append(f"已生效筛选：{applied}")
        text = "\n".join(lines)

        if chart and (income > 0 or expense > 0):
            if not self._ensure_matplotlib():
                return text + "\n（图表生成失败，仅返回文字）"
            png = os.path.join(get_astrbot_temp_path(), f"ha_stats_{uuid.uuid4().hex}.png")
            try:
                # 渲染是 CPU 密集，丢到线程池避免卡住事件循环
                await asyncio.to_thread(self._render_stats_png, data, str(applied or {}), png)
                caption = f"收支统计（单位：元）· 结余 {net / 100:,.2f} 元"
                await self.context.send_message(
                    event.unified_msg_origin,
                    MessageChain().message(caption).file_image(png),
                )
            except Exception as e:
                logger.error(f"[HomeAssistant] 发送统计图失败: {e}")
            finally:
                if os.path.exists(png):
                    os.remove(png)
        return text

    @filter.llm_tool(name="export_bills")
    async def export_bills(
        self,
        event: AstrMessageEvent,
        type: str = None,
        category: str = None,
        status: str = None,
        participant: str = None,
        month: str = None,
        year: int = None,
        from_date: str = None,
        to_date: str = None,
        amount_min: int = None,
        amount_max: int = None,
        account_id: int = None,
        destination: str = "text",
        relative_path: str = None,
        limit: int = 20,
    ) -> str:
        """查询账单明细（默认）或导出为 CSV 文件。纯查询模式返回明细文本供你直接回答，结果过长时只显示前 limit 条并提示导出；destination 可选 workspace（存当前会话工作区）/ user（直接发 CSV 文件）/ both。

        Args:
            type(string): income 或 expense（可选）。
            category(string): 分类筛选，如 餐饮（可选）。
            status(string): 状态筛选，settled 或 pending（可选）。
            participant(string): 参与人名字筛选（可选）。
            month(string): 月份 YYYY-MM（可选）。
            year(number): 年份（可选）。
            from_date(string): 起始日期 YYYY-MM-DD（可选）。
            to_date(string): 结束日期 YYYY-MM-DD（可选）。
            amount_min(number): 金额下限（单位：分，1 元 = 100 分）（可选）。
            amount_max(number): 金额上限（单位：分，1 元 = 100 分）（可选）。
            account_id(number): 记账账户 id（有多个账本时必须指定）。
            destination(string): text=纯查询返回明细文本（默认）；workspace=只存会话工作区；user=直接把 CSV 文件发给你；both=既存工作区又发文件。
            relative_path(string): destination 为 workspace/both 时，工作区内的相对路径，如 exports/家庭账单.csv（可选，默认自动命名）。
            limit(number): 纯查询模式下最多返回的明细条数，默认 20，最大 50（可选）。
        """
        filters = {}
        if type:
            filters["type"] = type
        if category:
            filters["category"] = category
        if status:
            filters["status"] = status
        if participant:
            filters["participant"] = participant
        if month:
            filters["month"] = month
        if year:
            filters["year"] = year
        if from_date:
            filters["from"] = from_date
        if to_date:
            filters["to"] = to_date
        if amount_min is not None:
            filters["amount_min"] = amount_min
        if amount_max is not None:
            filters["amount_max"] = amount_max
        if account_id:
            filters["account_id"] = account_id

        # ── 纯查询模式（默认）：返回明细文本给 BOT，不进 LLM 上下文超长风险 ──
        if destination == "text":
            qs = urlencode(filters)
            data = await self._api(
                "GET", f"api/v1/bills/export?{qs}", identity=self._identity(event)
            )
            if not data.get("ok", True):
                return self._err_msg(data)
            items = data.get("items") or []
            total = data.get("total") or len(items)
            applied = data.get("applied") or {}
            if not items:
                return "没有符合条件的账单。" + (f"\n已生效筛选：{applied}" if applied else "")
            try:
                cap = max(1, min(int(limit) if limit else 20, 50))
            except (TypeError, ValueError):
                cap = 20
            shown = items[:cap]
            lines = [f"共 {total} 条账单（显示前 {len(shown)} 条）："]
            for b in shown:
                kind = "收入" if b.get("type") == "income" else "支出"
                lines.append(
                    f"- #{b.get('id', '?')} {str(b.get('occurred_at', ''))[:10]} {kind} "
                    f"{b.get('category', '')} {b.get('amount', '')}分 "
                    f"{b.get('note', '') or ''}"
                )
            if total > len(shown):
                lines.append(
                    f"\n…还有 {total - len(shown)} 条未显示。"
                    f"可用 destination=user 把 CSV 文件发给你，或 destination=workspace 存工作区。"
                )
            if applied:
                lines.append(f"已生效筛选：{applied}")
            return "\n".join(lines)

        # ── 文件导出模式（workspace / user / both）：走 CSV ──
        params = dict(filters)
        params["format"] = "csv"
        qs = urlencode(params)

        ret = await self._api_raw(
            "GET", f"api/v1/bills/export?{qs}", identity=self._identity(event)
        )
        if isinstance(ret, dict):
            return self._err_msg(ret)
        http_status, csv_text = ret
        if http_status >= 400:
            return f"导出失败（HTTP {http_status}）：{csv_text[:300]}"
        if not isinstance(csv_text, str) or not csv_text.strip():
            return "导出失败：后端返回空内容。"
        try:
            # 用 csv.reader 精确统计行数（正确处理带引号换行的字段），不进 LLM 上下文
            row_count = max(0, sum(1 for _ in csv.reader(io.StringIO(csv_text))) - 1)
        except Exception:
            row_count = 0

        ws_path = None
        if destination in ("workspace", "both"):
            try:
                if relative_path:
                    target = self._resolve_workspace_rel(event, relative_path)
                else:
                    target = (
                        self._workspace_root(event)
                        / "exports"
                        / f"bills_{datetime.now():%Y%m%d_%H%M%S}.csv"
                    )
                target.parent.mkdir(parents=True, exist_ok=True)
                with open(target, "w", encoding="utf-8", newline="") as f:
                    f.write(csv_text)  # csv_text 自带 UTF-8 BOM
                ws_path = target.relative_to(self._workspace_root(event)).as_posix()
            except ValueError as e:
                return f"无效路径：{e}"

        if destination in ("user", "both"):
            filename = (
                os.path.basename(ws_path)
                if ws_path
                else f"bills_{datetime.now():%Y%m%d_%H%M%S}.csv"
            )
            tmp = os.path.join(get_astrbot_temp_path(), f"ha_export_{uuid.uuid4().hex}.csv")
            with open(tmp, "w", encoding="utf-8", newline="") as f:
                f.write(csv_text)
            try:
                ok = await self.context.send_message(
                    event.unified_msg_origin,
                    MessageChain([File(name=filename, file=tmp)]),
                )
                if not ok:
                    return f"文件已生成但发送失败（未匹配到平台）。工作区副本：{ws_path or filename}"
            finally:
                if os.path.exists(tmp):
                    os.remove(tmp)

        applied_note = f"筛选：{dict(params)}"
        if destination == "workspace":
            return (
                f"已导出 {row_count} 条账单到工作区：{ws_path}。\n"
                f"{applied_note}\n"
                f"可用文件工具读取/处理，或让我把它发给你。"
            )
        if destination == "user":
            return f"已把 {row_count} 条账单的 CSV 文件发给你。\n{applied_note}"
        return f"已导出 {row_count} 条账单到工作区：{ws_path}，并已把文件发给你。\n{applied_note}"

    # ────────────────────────── 文件导入（方向 B） ──────────────────────────

    def _parse_bills_file(self, path: str) -> tuple:
        """解析账单文件（xlsx/xls/csv）为 CreateBillInput 列表（宽容）。

        金额默认「元」→ 换算为「分」；类型/日期/参与人规范化；坏行跳过并记录原因。
        返回 (bills, skipped_count, reasons)。
        """
        bills = []
        skipped = 0
        reasons = []

        def norm_type(v):
            s = str(v).strip().lower() if v is not None else ""
            return {
                "收入": "income", "支出": "expense",
                "income": "income", "expense": "expense",
                "收": "income", "支": "expense",
            }.get(s)

        def norm_amount(v):
            if v is None:
                return None
            s = (
                str(v)
                .replace(",", "").replace("，", "").replace(" ", "")
                .replace("¥", "").replace("￥", "").replace("元", "")
                .strip()
            )
            try:
                num = float(s)
            except (TypeError, ValueError):
                return None
            fen = round(num * 100)  # 文件金额默认「元」→「分」
            return fen if fen >= 1 else None

        def norm_date(v):
            if v is None:
                return None
            if hasattr(v, "year"):  # openpyxl/xlrd 已解析为日期对象
                return f"{v.year:04d}-{v.month:02d}-{v.day:02d}"
            s = str(v).strip().replace("/", "-")
            m = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})", s)
            if m:
                return f"{int(m.group(1)):04d}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
            m = re.match(r"^(\d{1,2})-(\d{1,2})(?:[^0-9]|$)", s)  # MM-DD 补当年
            if m:
                return f"{_date.today().year:04d}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"
            return None

        def norm_category(v):
            cats = {"餐饮", "交通", "购物", "居住", "娱乐", "医疗", "教育", "人情", "其他"}
            s = str(v).strip() if v is not None else ""
            return s if s in cats else ""

        rows = []
        low = path.lower()
        try:
            if low.endswith(".csv"):
                text = None
                for enc in ("utf-8-sig", "gbk", "utf-8"):
                    try:
                        with open(path, "r", encoding=enc, newline="") as f:
                            text = f.read()
                        break
                    except (UnicodeDecodeError, UnicodeError):
                        continue
                if text is None:
                    return [], 0, ["无法识别文件编码"]
                rows = list(csv.reader(io.StringIO(text)))
            elif low.endswith(".xlsx"):
                import openpyxl

                wb = openpyxl.load_workbook(path, data_only=True)
                ws = wb[wb.sheetnames[0]]
                rows = [[c.value for c in row] for row in ws.iter_rows()]
            elif low.endswith(".xls"):
                import xlrd

                book = xlrd.open_workbook(path)
                sheet = book.sheet_by_index(0)
                rows = [sheet.row_values(i) for i in range(sheet.nrows)]
            else:
                return [], 0, [f"不支持的文件类型：{low.rsplit('.', 1)[-1]}"]
        except Exception as e:
            return [], 0, [f"文件读取失败：{e}"]

        # 表头识别（前 10 行找含关键词的表头行 → 列映射；找不到则按约定列序）
        header_keys = ("日期", "类型", "金额", "分类", "备注", "参与人")
        header_map = {}
        data_start = 0
        for i, row in enumerate(rows[:10]):
            joined = "".join(str(c) if c is not None else "" for c in row)
            if any(k in joined for k in header_keys):
                for idx, cell in enumerate(row):
                    s = str(cell) if cell is not None else ""
                    if ("日期" in s) or ("时间" in s):
                        header_map["date"] = idx
                    elif "类型" in s:
                        header_map["type"] = idx
                    elif ("金额" in s) or ("amount" in s.lower()):
                        header_map["amount"] = idx
                    elif "分类" in s:
                        header_map["category"] = idx
                    elif ("备注" in s) or ("说明" in s):
                        header_map["note"] = idx
                    elif ("参与人" in s) or ("成员" in s):
                        header_map["participants"] = idx
                data_start = i + 1
                break

        for i, row in enumerate(rows[data_start:]):
            cells = [c if c is not None else "" for c in row]
            if all(str(c).strip() == "" for c in cells):
                continue  # 空行
            lineno = i + data_start + 1
            if header_map:
                def getc(k):
                    idx = header_map.get(k)
                    return cells[idx] if idx is not None and idx < len(cells) else None

                t = norm_type(getc("type"))
                amt = norm_amount(getc("amount"))
                d = norm_date(getc("date"))
                cat = norm_category(getc("category"))
                note = str(getc("note") or "").strip()
                parts = getc("participants")
            else:  # 约定列序：日期 类型 金额 分类 备注 参与人
                t = norm_type(cells[1] if len(cells) > 1 else "")
                amt = norm_amount(cells[2] if len(cells) > 2 else "")
                d = norm_date(cells[0] if len(cells) > 0 else "")
                cat = norm_category(cells[3] if len(cells) > 3 else "")
                note = str(cells[4] if len(cells) > 4 else "").strip()
                parts = cells[5] if len(cells) > 5 else None
            if not t:
                skipped += 1
                reasons.append(f"第 {lineno} 行：类型无法识别")
                continue
            if amt is None:
                skipped += 1
                reasons.append(f"第 {lineno} 行：金额无法识别")
                continue
            bill = {"type": t, "amount": amt}
            if d:
                bill["occurred_at"] = d
            if cat:
                bill["category"] = cat
            if note:
                bill["note"] = note
            if parts:
                plist = self._normalize_participants(
                    [p.strip() for p in str(parts).replace("、", ",").split(",") if p.strip()]
                )
                if plist:
                    bill["participants"] = plist
            bills.append(bill)
        return bills, skipped, reasons

    @filter.llm_tool(name="save_uploaded_file")
    async def save_uploaded_file(self, event: AstrMessageEvent, relative_path: str = None) -> str:
        """把用户发送的文件保存到当前会话工作区（只处理第一个文件附件）。保存后可 parse_bills_file 预览或 import_bills 导入记账。

        Args:
            relative_path(string): 工作区内的相对路径，如 bills/对账单.xlsx（可选，默认用原文件名）。
        """
        file_comp = None
        for comp in event.message_obj.message:
            if isinstance(comp, File):
                file_comp = comp
                break
            if isinstance(comp, Reply) and getattr(comp, "chain", None):
                for rc in comp.chain:
                    if isinstance(rc, File):
                        file_comp = rc
                        break
                if file_comp:
                    break
        if not file_comp:
            return "未检测到文件，请先发送文件附件。"
        src = await file_comp.get_file()  # 异步获取，勿用 comp.file
        if not src:
            return "无法获取文件内容，请重新发送。"
        try:
            if relative_path:
                target = self._resolve_workspace_rel(event, relative_path)
            else:
                name = file_comp.name or os.path.basename(src) or "uploaded.bin"
                target = self._workspace_root(event) / name
            target.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy(src, target)
        except ValueError as e:
            return f"无效路径：{e}"
        finally:
            if src and os.path.exists(src):
                os.remove(src)  # AstrBot 不自动清理 temp 副本，必须手动删
        rel = target.relative_to(self._workspace_root(event)).as_posix()
        size = os.path.getsize(target)
        return f"已保存文件到工作区：{rel}（{size} 字节）。可 parse_bills_file 预览，或 import_bills 直接导入记账。"

    @filter.llm_tool(name="parse_bills_file")
    async def parse_bills_file(self, event: AstrMessageEvent, relative_path: str, limit: int = 20) -> str:
        """解析工作区里的账单文件（xlsx/xls/csv）为结构化账单行，返回明细文本供确认。文件金额默认「元」，已换算为「分」展示。

        Args:
            relative_path(string): 工作区内的相对路径，如 bills/对账单.xlsx。
            limit(number): 最多显示的明细条数，默认 20，最大 50（可选）。
        """
        try:
            target = self._resolve_workspace_rel(event, relative_path)
        except ValueError as e:
            return f"无效路径：{e}"
        if not target.exists() or not target.is_file():
            return f"文件不存在：{relative_path}"
        bills, skipped, reasons = self._parse_bills_file(str(target))
        if not bills:
            tail = "；".join(reasons[:5]) if reasons else "无有效行"
            return f"未能从文件解析出有效账单（跳过 {skipped} 行）。{tail}"
        try:
            cap = max(1, min(int(limit) if limit else 20, 50))
        except (TypeError, ValueError):
            cap = 20
        lines = [f"共 {len(bills)} 条有效账单（显示前 {min(cap, len(bills))} 条）："]
        for b in bills[:cap]:
            kind = "收入" if b.get("type") == "income" else "支出"
            parts_txt = f" 参与人:{','.join(p.get('name', '') for p in b.get('participants', []))}" if b.get("participants") else ""
            lines.append(
                f"- {b.get('occurred_at', '')} {kind} {b.get('category', '')} "
                f"{b.get('amount', 0)}分 {b.get('note', '')}{parts_txt}"
            )
        extra = []
        if len(bills) > cap:
            extra.append(f"…还有 {len(bills) - cap} 条，可 import_bills 直接导入。")
        if skipped:
            extra.append(f"跳过 {skipped} 行坏数据：" + "；".join(reasons[:3]) + "…")
        if extra:
            lines.append("\n".join(extra))
        return "\n".join(lines)

    @filter.llm_tool(name="import_bills")
    async def import_bills(self, event: AstrMessageEvent, relative_path: str, account_id: int = None) -> str:
        """解析工作区里的账单文件并批量导入记账（宽容：坏行跳过、合法行入库，返回失败明细）。

        Args:
            relative_path(string): 工作区内的相对路径，如 bills/对账单.xlsx。
            account_id(number): 记账账户 id（有多个账本时必须指定）。
        """
        try:
            target = self._resolve_workspace_rel(event, relative_path)
        except ValueError as e:
            return f"无效路径：{e}"
        if not target.exists() or not target.is_file():
            return f"文件不存在：{relative_path}"
        bills, skipped, reasons = self._parse_bills_file(str(target))
        if not bills:
            tail = "；".join(reasons[:5]) if reasons else "无有效行"
            return f"文件中没有可导入的账单（跳过 {skipped} 行）。{tail}"
        if len(bills) > 500:
            return f"文件含 {len(bills)} 条，超过单次 500 条上限，请拆分后再导入。"
        body = {"bills": bills}
        if account_id:
            body["account_id"] = account_id
        data = await self._api(
            "POST", "api/v1/bills/import",
            identity=self._identity(event), json_body=body,
        )
        if not data.get("ok", True):
            return self._err_msg(data)
        ok_count = data.get("ok_count", 0)
        fail_count = data.get("fail_count", 0)
        total = data.get("total", len(bills))
        lines = [f"已导入 {ok_count}/{total} 笔账单。"]
        if fail_count:
            lines.append(f"失败 {fail_count} 笔：")
            for f in (data.get("failures") or [])[:10]:
                lines.append(
                    f"- 第 {f.get('index', '?')} 行：{f.get('code', '')} {f.get('message', '')}"
                )
            if fail_count > 10:
                lines.append(f"…共 {fail_count} 笔失败")
        if skipped:
            lines.append(f"（文件解析跳过 {skipped} 行：" + "；".join(reasons[:3]) + "…）")
        return "\n".join(lines)

    @filter.llm_tool(name="delete_file")
    async def delete_file(self, event: AstrMessageEvent, relative_path: str) -> str:
        """删除当前会话工作区里的一个文件或目录（不可恢复，请谨慎）。

        Args:
            relative_path(string): 工作区内的相对路径。
        """
        try:
            target = self._resolve_workspace_rel(event, relative_path)
        except ValueError as e:
            return f"无效路径：{e}"
        if not target.exists():
            return f"文件不存在：{relative_path}"
        try:
            if target.is_dir():
                shutil.rmtree(target)
            else:
                os.remove(target)
        except OSError as e:
            return f"删除失败：{e}"
        logger.info(f"[HomeAssistant] 已删除工作区文件: {relative_path}")
        return f"已删除：{relative_path}"

    # ────────────────────────── outbox 轮询 ──────────────────────────

    async def _outbox_loop(self) -> None:
        while True:
            try:
                await self._poll_outbox()
            except Exception as e:
                logger.error(f"[HomeAssistant] outbox 轮询出错: {e}")
            await asyncio.sleep(self.poll_interval)

    async def _poll_outbox(self) -> None:
        if not self.api_key:
            return
        data = await self._api("GET", "api/v1/outbox/pending?channel=qq&limit=10")
        for item in data.get("items") or []:
            await self._deliver_outbox(item)

    async def _deliver_outbox(self, row: dict) -> None:
        oid = row.get("id")
        target = row.get("target_id") or ""
        content = row.get("content") or ""
        if not oid or not target or not content:
            return
        try:
            umo = self._session_map.get(target) or self._build_umo(target)
            if not umo:
                logger.warning(
                    f"[HomeAssistant] outbox#{oid} 无法定位会话（target={target}）"
                )
                await self._receipt(oid, "failed", "no session for target")
                return
            ok = await self.context.send_message(umo, MessageChain().message(content))
            if ok:
                logger.info(f"[HomeAssistant] outbox#{oid} 已投递到 QQ")
                await self._receipt(oid, "sent")
            else:
                logger.warning(f"[HomeAssistant] outbox#{oid} send_message 未匹配到平台")
                await self._receipt(oid, "failed", "send_message no platform matched")
        except Exception as e:
            logger.error(f"[HomeAssistant] outbox#{oid} 投递异常: {e}")
            await self._receipt(oid, "failed", str(e)[:300])

    def _build_umo(self, openid: str) -> str | None:
        """无注册表命中时按 QQ 官方 C2C 格式构造会话：platform_id:FriendMessage:openid"""
        if not self._platform_id or not openid:
            return None
        return f"{self._platform_id}:FriendMessage:{openid}"

    async def _receipt(self, oid: int, status: str, error: str = None) -> None:
        body = {"status": status}
        if error:
            body["error"] = error
        await self._api("POST", f"api/v1/outbox/{oid}/delivery", json_body=body)
